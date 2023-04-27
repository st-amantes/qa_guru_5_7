import os.path
import zipfile
import requests
import xlrd
from selenium import webdriver
import time
from pypdf import PdfReader
from openpyxl import load_workbook
import csv
from selene import browser
from zipfile import ZipFile

Current_file_path = os.path.abspath(__file__)
Folder_path = os.path.dirname(__file__)
Resource_path = os.path.join(Folder_path, 'resources')
tmp_path = os.path.join(Folder_path, 'tmp')


def test_csv():
    csv_path = os.path.join(Resource_path, 'age.csv')
    with open(csv_path, 'w') as csv_file:
        csvwriter = csv.writer(csv_file, delimiter=',')
        csvwriter.writerow(['Boris', 'Sergey', 'Viktor'])
        csvwriter.writerow(['Mihail', 'Evgen', 'Makar'])

    with open(csv_path) as csv_file:
        csvreader = csv.reader(csv_file)
        name = []
        for row in csvreader:
            name.append(row)
            print(row)
    assert name[0] == ['Boris', 'Sergey', 'Viktor']


def test_pdf():
    pdf_path = os.path.join(Resource_path, 'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(pdf_path)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    print(page)
    print(number_of_pages)
    print(text)
    assert number_of_pages == 412


def test_xls():
    path_xls = os.path.join(Resource_path, 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(path_xls)
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')

    for rx in range(sheet.nrows):
        print(sheet.row(rx))

    assert book.nsheets == 1
    assert book.sheet_names() == ['Sheet1']
    assert sheet.ncols == 8
    assert sheet.nrows == 10
    assert sheet.cell_value(rowx=0, colx=1) == 'First Name'


def test_xlsx():
    path_xlsx = os.path.join(Resource_path, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(path_xlsx)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'


def test_download_file():
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'

    tmp_folder = os.path.join(Folder_path, 'tmp', 'selenium_logo.png')
    if not os.path.exists(tmp_folder):
        os.mkdir('tmp')

    r = requests.get(url)
    with open(tmp_folder, 'wb') as file:
        file.write(r.content)
    size_file = os.path.getsize(tmp_folder)
    assert size_file == 30803


def test_download_file_with_browser():
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": tmp_path,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    browser.config.driver_options = options

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(10)

    assert os.path.exists(os.path.join(tmp_path, 'pytest-main.zip'))
    assert os.path.getsize(os.path.join(tmp_path, 'pytest-main.zip')) > 0


def test_add_zip_file():
    Project_path = os.path.dirname(os.path.abspath(__file__))
    zip_path = os.path.join(Project_path, 'resources')
    zip_file = os.path.join('resources.zip')
    with zipfile.ZipFile(zip_file, 'a') as zip_files:
        for i in os.listdir(zip_path):
            file_path = os.path.join(zip_path, i)
            zip_files.write(file_path, i)

    with zipfile.ZipFile(zip_file, "r") as zip_files:
        for i in os.listdir(zip_path):
            assert i in zip_files.namelist()
